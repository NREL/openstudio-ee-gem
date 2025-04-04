from pathlib import Path
from openpyxl import load_workbook

# Define current and target paths
CURRENT_DIR_PATH = Path(__file__).parent.absolute()
excel_path = CURRENT_DIR_PATH.parent / 'resources' / 'optimization.xlsx'
new_excel_output_path = CURRENT_DIR_PATH.parent / 'resources' / 'optimization_updated.xlsx'

# Load workbook
wb = load_workbook(excel_path, data_only=False, keep_links=True)

# Select the 'values' worksheet specifically
if 'values' not in wb.sheetnames:
    raise ValueError("Sheet named 'values' not found in the Excel file.")
ws = wb['values']

# Find the row with "Embodied Carbon" in the first column
target_row = None
for row in ws.iter_rows(min_row=1, max_col=1):
    cell = row[0]
    if str(cell.value).strip() == 'Embodied Carbon':
        target_row = cell.row
        break

# Find the column with "Scenario_1" in the header row
target_col = None
for cell in ws[1]:  # First row is assumed to be the header
    if str(cell.value).strip() == 'Scenario_1':
        target_col = cell.column
        break

# Modify the value if both row and column are found
if target_row and target_col:
    ws.cell(row=target_row, column=target_col).value = 10
else:
    raise ValueError("Could not find 'Embodied Carbon' row or 'Scenario_1' column.")

# Save updated workbook
wb.save(new_excel_output_path)


