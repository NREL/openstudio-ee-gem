# *******************************************************************************
# OpenStudio(R), Copyright (c) Alliance for Sustainable Energy, LLC.
# See also https://openstudio.net/license
# *******************************************************************************

import openstudio
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
import plotly.graph_objects as go
from openpyxl import load_workbook

CURRENT_DIR_PATH = Path(__file__).absolute()
optimization_excel_path = CURRENT_DIR_PATH.parent / 'resources' / 'optimization.xlsx'
new_optimization_excel_output_path = CURRENT_DIR_PATH.parent / 'resources' / 'optimization_updated.xlsx'

def modify_optimization_sheet(replacement_value):

    """Reads and modifies Excel spreadsheet for optimization visualizations"""
    wb = load_workbook(optimization_excel_path, data_only=False, keep_links=True)

    # Select the 'values' worksheet specifically
    if 'values' not in wb.sheetnames:
        raise ValueError("Sheet named 'values' not found in the Excel file.")
    ws = wb['values']

    # Find the row with "Embodied Carbon" in the first column
    target_row = None
    for row in ws.iter_rows(min_row=1, max_col=1):
        cell = row[0]
        if str(cell.value).strip() == 'Embodied Carbon':
            target_row = cell.row
            break

    # Find the column with "Scenario_1" in the header row
    target_col = None
    for cell in ws[1]:  # First row is assumed to be the header
        if str(cell.value).strip() == 'Scenario_1':
            target_col = cell.column
            break

    # Modify the value if both row and column are found
    if target_row and target_col:
        ws.cell(row=target_row, column=target_col).value = replacement_value
    else:
        raise ValueError("Could not find 'Embodied Carbon' row or 'Scenario_1' column.")

    # Save updated workbook
    wb.save(new_optimization_excel_output_path)

def optimization(self):

    optimization_weights = pd.read_excel(optimization_excel_path, sheet_name = "weights") # Not being currently used
    factor_values = pd.read_excel(optimization_excel_path, sheet_name = "values")
    n_scenarios = 3

    for scenario in range(1,n_scenarios+1):
        #print(scenario)
        factor_values["Normalized_Scenario_" + str(scenario)] = factor_values["Scenario_" + str(scenario)]/factor_values["Basis"]

    fig = go.Figure()

    for scenario in range(1, n_scenarios+1):
        fig.add_trace(
            go.Scatterpolar(
                theta = factor_values["Factor"],
                r = factor_values["Normalized_Scenario_" + str (scenario)], name = "Scenario_" + str(scenario) 
            ))

    fig.show()

class ECReport(openstudio.measure.ReportingMeasure):
    def name(self):
        return "ReportAdditionalProperties"

    def description(self):
        return "Reports all AdditionalProperties objects and their key-value pairs in the model."

    def modeler_description(self):
        return "Traverses the model and extracts data from AdditionalProperties objects."

    def __init__(self):
        super().__init__()
        self.material_data = {}

    def parse_workspace_objects(self, objects):
        """Processes a list of OS:AdditionalProperties objects and returns the total numeric value."""
        total = 0.0

        for obj in objects:
            num_fields = obj.numFields()
            if num_fields < 5:
                continue

            material_name = obj.getString(1, True)
            numeric_value = obj.getString(num_fields - 1, True)

            if material_name.is_initialized() and numeric_value.is_initialized():
                try:
                    value = float(numeric_value.get())
                    self.material_data[material_name.get()] = value
                    total += value
                except ValueError:
                    print(f"Warning: Could not convert {numeric_value.get()} to float for {material_name.get()}")

        return round(total, 2)

    def run(self, runner, model):
        self.material_data.clear()

        additional_properties_objects = model.getObjectsByType(openstudio.IddObjectType("OS_AdditionalProperties"))

        if additional_properties_objects:
            runner.registerInfo(f"Found {len(additional_properties_objects)} AdditionalProperties objects.")
            total_gwp = self.parse_workspace_objects(additional_properties_objects)
        else:
            total_gwp = 0.0
        print(f"Total GWP value = {total_gwp}")
        runner.registerInfo(f"Extracted Material Data:{str(self.material_data)}")

        # Write result to Excel
        modify_optimization_sheet(total_gwp)
        optimization(self)

        runner.registerInfo("Cleaning up model from memory.")
        del additional_properties_objects

        del model

        return True

# Register the measure
measure = ECReport()
